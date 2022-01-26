""""
Description: Check if IP address(es) is reachable AND store the results in an excel file
Date:       24-Jan-2022
Version:    1.0
"""
__author__ = 'Madhava Srinivas A'

import nmap
from openpyxl import Workbook

wb = Workbook()
wb.create_sheet(title='PSLabIPList')
wb.active = wb.sheetnames.index('PSLabIPList')
ws = wb.active
ws.append(['IP ADDRESS', 'STATE'])
wb.save(filename='IPAM.xlsx')


def checkip(ips: str) -> None:
    scanner = nmap.PortScanner()
    # host = socket.gethostbyaddr(ip)
    scanner.scan(hosts=ips, arguments='-v')
    hosts_list = [(host, scanner[host]['status']['state']) for host in scanner.all_hosts()]
    for item in hosts_list:
        print(item)
        ws.append(item)


checkip('16.32.12.50-250')
wb.save(filename='IPAM.xlsx')